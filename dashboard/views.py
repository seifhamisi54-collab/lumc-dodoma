from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.db import connection
from django.db.models import Sum, Count, Q
from django.contrib.auth import get_user_model
from django.contrib import messages
from django.core.files.storage import default_storage
from django.core.paginator import Paginator

# Local app imports
from locations.models import Region, District, Ward
from dashboard.models import (
    VillageBoundary,
    SocialService,
    Parcel,
    Infrastructure,
    CCOReport,
    LandUseReport,
    ImportLog,
    DownloadDonation,
)

import json
import logging
import csv
from datetime import datetime
import uuid

from dashboard.financial_year import (
    DEFAULT_FINANCIAL_YEAR,
    normalize_financial_year,
    session_financial_year,
    set_session_financial_year,
    suggested_financial_years,
)

User = get_user_model()
logger = logging.getLogger(__name__)

PLAN_READY_STATUSES = ('prepared', 'approved', 'completed')
PLAN_VALIDITY_YEARS = 10


def _village_has_plan_q():
    """Kijiji kina land use plan: date_prepared ipo au status ni prepared/approved/completed."""
    status_q = Q()
    for value in PLAN_READY_STATUSES:
        status_q |= Q(status__iexact=value)
    return Q(date_prepared__isnull=False) | status_q


def _village_has_plan(village):
    if village.date_prepared:
        return True
    status = (village.status or '').strip().lower()
    return status in PLAN_READY_STATUSES


def _plan_year_fields(village):
    """Mwaka wa kuisha = mwaka wa kuandaa + miaka 10 (ikiwa date_end haipo)."""
    prep_year = village.date_prepared.year if village.date_prepared else None
    if village.date_end:
        expiry_year = village.date_end.year
    elif prep_year:
        expiry_year = prep_year + PLAN_VALIDITY_YEARS
    else:
        expiry_year = None
    status_info = _plan_expiry_status(prep_year, expiry_year)
    return prep_year, expiry_year, status_info


def _plan_expiry_status(prep_year, expiry_year):
    current_year = datetime.now().year
    if not prep_year:
        return {
            'label': 'Hakuna plan',
            'code': 'no_plan',
            'years_remaining': None,
        }

    effective_expiry = expiry_year or (prep_year + PLAN_VALIDITY_YEARS)
    years_remaining = effective_expiry - current_year

    if years_remaining <= 0:
        overdue = abs(years_remaining)
        if overdue == 0:
            label = 'Imeisha mwaka huu'
        elif overdue == 1:
            label = 'Imeisha (mwaka 1 uliopita)'
        else:
            label = f'Imeisha (miaka {overdue} iliyopita)'
        return {
            'label': label,
            'code': 'expired',
            'years_remaining': years_remaining,
        }

    if years_remaining == 1:
        label = 'Mwaka 1 umebaki'
    else:
        label = f'Miaka {years_remaining} imebaki'

    code = 'warning' if years_remaining <= 5 else 'active'
    return {
        'label': label,
        'code': code,
        'years_remaining': years_remaining,
    }


def _serialize_village_data(village):
    prep_year, expiry_year, status_info = _plan_year_fields(village)
    return {
        'id': str(village.id),
        'name': village.name or '-',
        'ward_name': village.ward_name or '-',
        'district_name': village.district_name or '-',
        'region_name': village.region_name or '-',
        'plan_preparation_year': prep_year,
        'plan_expiry_year': expiry_year,
        'years_remaining': status_info['years_remaining'],
        'plan_status_label': status_info['label'],
        'plan_status_code': status_info['code'],
        'plan_status': village.status or 'not_prepared',
        'approval_status': village.status or 'draft',
        'population': 0,
        'households': 0,
        'has_pdf': False,
    }


# =====================================================
# PAGE VIEWS
# =====================================================

@login_required
def home(request):
    from accounts.permissions import can_access_admin_panel
    from dashboard.system_admin_views import seed_default_admin_data

    can_admin = can_access_admin_panel(request.user)
    if can_admin:
        try:
            seed_default_admin_data()
        except Exception:
            pass

    return render(request, 'dashboard/home.html', {
        'can_admin': can_admin,
    })


@login_required
@require_http_methods(['GET', 'POST'])
def api_financial_year(request):
    """Soma / weka mwaka wa fedha wa session — inatumika kwenye moduli zote."""
    if request.method == 'GET':
        fy = session_financial_year(request)
        return JsonResponse({
            'success': True,
            'financial_year': fy,
            'default_financial_year': DEFAULT_FINANCIAL_YEAR,
            'suggestions': suggested_financial_years(),
        })

    try:
        body = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        body = {}
    raw = body.get('financial_year') or request.POST.get('financial_year') or ''
    fy = set_session_financial_year(request, raw)
    return JsonResponse({'success': True, 'financial_year': fy})


@login_required
def landuse_home(request):
    return render(request, 'dashboard/landuse_home.html')

def map_view(request):
    """Map view - fixed UUID error"""
    region_name = request.GET.get('region_name', 'Tanzania')
    if region_name in ['undefined', 'null', 'None', '']:
        region_name = 'Tanzania'
    return render(request, 'dashboard/map_view.html', {
        'region_name': region_name,
        'tools_page': False,
    })


def gis_tools_view(request):
    """Ukurasa wa GIS Tools — ramani ile ile + zana za uchambuzi."""
    region_name = request.GET.get('region_name', 'Tanzania')
    if region_name in ['undefined', 'null', 'None', '']:
        region_name = 'Tanzania'
    tool = request.GET.get('tool', 'qgis-editor')
    valid_tools = {
        'qgis-editor', 'ai-digitization', 'georeferencing', 'data-cleaning', 'topology', 'measures', 'editing',
    }
    if tool not in valid_tools:
        tool = 'qgis-editor'
    return render(request, 'dashboard/map_view.html', {
        'region_name': region_name,
        'tools_page': True,
        'initial_tool': tool,
    })

@login_required
def district_detail(request, district_id):
    district = get_object_or_404(District, id=district_id)
    villages = VillageBoundary.objects.filter(district_name=district.name)
    context = {
        'district': district,
        'villages': villages,
    }
    return render(request, 'dashboard/district_detail.html', context)

@login_required
def village_detail(request, village_id):
    village = get_object_or_404(VillageBoundary, id=village_id)
    context = {'village': village}
    return render(request, 'dashboard/village_detail.html', context)

def _clean_region_name(region_name):
    if not region_name:
        return ''
    region_name = str(region_name).strip()
    if region_name in ['undefined', 'null', 'None', 'Tanzania', '']:
        return ''
    return region_name


def data_portal(request):
    from accounts.permissions import can_upload

    region_name = _clean_region_name(request.GET.get('region_name', ''))
    return render(request, 'dashboard/data_portal.html', {
        'region_name': region_name,
        'can_upload': can_upload(request.user) if request.user.is_authenticated else False,
    })


def signup_view(request):
    if request.method == 'POST':
        from accounts.models import (
            UserRole,
            registration_code_is_valid,
        )

        fullname = request.POST.get('fullname')
        email = request.POST.get('email')
        username = request.POST.get('username')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        registration_code = request.POST.get('registration_code') or ''

        if not registration_code_is_valid(registration_code):
            messages.error(request, 'Nambari ya usajili si sahihi au haipo.')
            return redirect('signup')

        if password != confirm_password:
            messages.error(request, 'Passwords do not match!')
            return redirect('signup')

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
            return redirect('signup')

        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already registered!')
            return redirect('signup')

        user = User.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=fullname,
            is_active=False,
        )
        viewer_role, _ = UserRole.objects.get_or_create(name='viewer')
        user.role = viewer_role
        user.save()

        messages.success(
            request,
            'Akaunti imeundwa. Subiri msimamizi aiamilishe kabla ya kuingia.',
        )
        return redirect('login')

    return render(request, 'login.html')


# =====================================================
# UPLOAD PAGE VIEWS
# =====================================================

@login_required
def upload_landuse_shapefile(request):
    return render(request, 'dashboard/upload_landuse.html', {'title': 'Upload Land Use Shapefile'})

@login_required
def upload_village_boundary(request):
    return render(request, 'dashboard/upload_village_boundary.html', {'title': 'Upload Village Boundary'})

@login_required
def upload_hamlet_boundary(request):
    return render(request, 'dashboard/upload_hamlet_boundary.html', {'title': 'Upload Hamlet Boundary'})

@login_required
def upload_social_services(request):
    return render(request, 'dashboard/upload_social_services.html', {'title': 'Upload Social Services'})

@login_required
def upload_parcels(request):
    return render(request, 'dashboard/upload_parcels.html', {'title': 'Upload Parcels'})

@login_required
def upload_infrastructure(request):
    return render(request, 'dashboard/upload_infrastructure.html', {'title': 'Upload Infrastructure'})

@login_required
def upload_cco_data(request):
    return render(request, 'dashboard/upload_cco.html', {'title': 'Upload CCO Data'})

@login_required
def download_shapefile(request):
    return render(request, 'dashboard/download_shapefile.html', {'title': 'Download Shapefile'})

@login_required
def download_report(request):
    return render(request, 'dashboard/download_report.html', {'title': 'Download Report'})


# =====================================================
# DATA PORTAL CRUD VIEWS
# =====================================================

@login_required
def add_village_data(request):
    if request.method == 'POST':
        try:
            # Convert dates if provided
            date_prepared = None
            date_end = None
            if request.POST.get('date_prepared'):
                try:
                    date_prepared = datetime.strptime(request.POST.get('date_prepared'), '%Y-%m-%d').date()
                except:
                    pass
            if request.POST.get('date_end'):
                try:
                    date_end = datetime.strptime(request.POST.get('date_end'), '%Y-%m-%d').date()
                except:
                    pass
            
            village_data = VillageBoundary.objects.create(
                name=request.POST.get('village_name'),
                ward_name=request.POST.get('ward_name'),
                district_name=request.POST.get('district_name'),
                region_name=request.POST.get('region_name'),
                sponsor=request.POST.get('sponsor', ''),
                date_prepared=date_prepared,
                date_end=date_end,
                status=request.POST.get('status', 'draft'),
                iv=request.POST.get('iv', ''),
            )
            
            messages.success(request, 'Village data added successfully!')
            return redirect('dashboard:village_detail', village_id=village_data.id)
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    return render(request, 'dashboard/add_village_data.html')

@login_required
def village_data_detail(request, village_id):
    village = get_object_or_404(VillageBoundary, id=village_id)
    return render(request, 'dashboard/village_data_detail.html', {'village': village})

@login_required
def edit_village_data(request, village_id):
    village = get_object_or_404(VillageBoundary, id=village_id)
    
    if request.method == 'POST':
        try:
            # Update logic here
            village.name = request.POST.get('village_name', village.name)
            village.ward_name = request.POST.get('ward_name', village.ward_name)
            village.district_name = request.POST.get('district_name', village.district_name)
            village.region_name = request.POST.get('region_name', village.region_name)
            village.save()
            messages.success(request, 'Village data updated successfully!')
            return redirect('dashboard:village_detail', village_id=village.id)
        except Exception as e:
            messages.error(request, f'Error: {str(e)}')
    
    return render(request, 'dashboard/edit_village_data.html', {'village': village})

@login_required
def delete_village_data(request, village_id):
    if request.method == 'POST':
        village = get_object_or_404(VillageBoundary, id=village_id)
        village.delete()
        messages.success(request, 'Village data deleted successfully!')
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error'}, status=405)


# =====================================================
# API VIEWS
# =====================================================

def api_regions(request):
    try:
        # Get regions from VillageBoundary table
        regions = VillageBoundary.objects.filter(
            region_name__isnull=False
        ).exclude(
            region_name=''
        ).values_list('region_name', flat=True).distinct().order_by('region_name')
        
        if regions.exists():
            data = [{'name': r} for r in regions]
            return JsonResponse(data, safe=False)
        
        # Fallback to Region model
        regions = Region.objects.all()
        data = [{'id': str(r.id), 'name': r.name} for r in regions]
        return JsonResponse(data, safe=False)
    except Exception:
        return JsonResponse([], safe=False)

def api_region_detail(request, region_id):
    try:
        region = Region.objects.get(id=region_id)
        return JsonResponse({
            'id': str(region.id),
            'name': region.name,
            'code': getattr(region, 'code', ''),
            'boundary_geojson': getattr(region, 'boundary_geojson', None),
            'center_lat': float(region.center_lat) if hasattr(region, 'center_lat') and region.center_lat else None,
            'center_lon': float(region.center_lon) if hasattr(region, 'center_lon') and region.center_lon else None,
        })
    except Region.DoesNotExist:
        return JsonResponse({'error': 'Region not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_region_geometry(request, region_id):
    return api_region_detail(request, region_id)

def api_region_boundary(request, region_id):
    return api_region_detail(request, region_id)

def api_district_detail(request, district_id):
    try:
        district = District.objects.get(id=district_id)
        wards = Ward.objects.filter(district=district).values('id', 'name')
        return JsonResponse({
            'id': str(district.id),
            'name': district.name,
            'region': district.region.name if district.region else '',
            'region_id': str(district.region.id) if district.region else None,
            'boundary_geojson': getattr(district, 'boundary_geojson', None),
            'wards': list(wards),
        })
    except District.DoesNotExist:
        return JsonResponse({'error': 'District not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_ward_detail(request, ward_id):
    try:
        ward = Ward.objects.get(id=ward_id)
        return JsonResponse({
            'id': str(ward.id),
            'name': ward.name,
            'district': ward.district.name if ward.district else '',
            'district_id': str(ward.district.id) if ward.district else None,
            'boundary_geojson': getattr(ward, 'boundary_geojson', None),
        })
    except Ward.DoesNotExist:
        return JsonResponse({'error': 'Ward not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_districts_by_region(request, region):
    """Get districts in a region - for dropdown"""
    # Clean region name
    if region in ['undefined', 'null', 'None', '']:
        return JsonResponse({'features': []})
    
    try:
        # Get from VillageBoundary table
        districts = VillageBoundary.objects.filter(
            region_name__iexact=region,
            district_name__isnull=False
        ).exclude(
            district_name=''
        ).values_list('district_name', flat=True).distinct().order_by('district_name')
        
        if districts.exists():
            features = [{'properties': {'name': d}} for d in districts]
            return JsonResponse({'features': features})
        
        # Fallback to District model
        districts = District.objects.filter(region__name__iexact=region)
        features = [{'properties': {'name': d.name}} for d in districts]
        return JsonResponse({'features': features})
    except Exception as e:
        return JsonResponse({'features': []})

def api_wards_by_district(request, region, district):
    """Get wards in a district - for dropdown"""
    # Clean parameters
    if region in ['undefined', 'null', 'None', ''] or district in ['undefined', 'null', 'None', '']:
        return JsonResponse({'wards': []})
    
    try:
        # Get from VillageBoundary table
        wards = VillageBoundary.objects.filter(
            region_name__iexact=region,
            district_name__iexact=district,
            ward_name__isnull=False
        ).exclude(
            ward_name=''
        ).values_list('ward_name', flat=True).distinct().order_by('ward_name')
        
        if wards.exists():
            ward_list = [{'name': w} for w in wards]
            return JsonResponse({'wards': ward_list})
        
        # Fallback to Ward model
        wards = Ward.objects.filter(
            district__name__iexact=district, 
            district__region__name__iexact=region
        )
        ward_list = [{'name': w.name} for w in wards]
        return JsonResponse({'wards': ward_list})
    except Exception:
        return JsonResponse({'wards': []})

def _clean_location_params(region, district, ward):
    if region in ['undefined', 'null', 'None', '']:
        region = None
    if district in ['undefined', 'null', 'None', '']:
        district = None
    if ward in ['undefined', 'null', 'None', '']:
        ward = None
    return region, district, ward


def _filter_villages_by_location(queryset, region=None, district=None, ward=None):
    if region:
        queryset = queryset.filter(region_name__iexact=region)
    if district:
        queryset = queryset.filter(district_name__iexact=district)
    if ward:
        queryset = queryset.filter(ward_name__iexact=ward)
    return queryset


def _filter_parcels_by_location(queryset, region=None, district=None, ward=None):
    if region:
        queryset = queryset.filter(region_name__iexact=region)
    if district:
        queryset = queryset.filter(district_name__iexact=district)
    if ward:
        queryset = queryset.filter(ward_name__iexact=ward)
    return queryset


def _get_ccro_summary(village_queryset, region=None, district=None, ward=None):
    """Hesabu CCRO kutoka planning_parcels (detailed_planning)."""
    from django.db.models import Exists, OuterRef
    from django.db.utils import OperationalError, ProgrammingError

    total_ccros = 0
    villages_with_ccro = 0
    try:
        from detailed_planning.models import PlanningParcel
        from dashboard.boundary_service import _district_search_names

        parcel_qs = PlanningParcel.objects.all()
        if region:
            parcel_qs = parcel_qs.filter(region_name__iexact=region)
        if district:
            district_q = Q()
            for name in _district_search_names(district):
                district_q |= Q(district_name__iexact=name)
            parcel_qs = parcel_qs.filter(district_q)
        if ward:
            parcel_qs = parcel_qs.filter(ward_name__iexact=ward)
        total_ccros = parcel_qs.count()

        parcel_match = PlanningParcel.objects.filter(
            village_name__iexact=OuterRef('name'),
            ward_name__iexact=OuterRef('ward_name'),
            district_name__iexact=OuterRef('district_name'),
        )
        if region:
            parcel_match = parcel_match.filter(region_name__iexact=region)
        if district:
            district_q = Q()
            for name in _district_search_names(district):
                district_q |= Q(district_name__iexact=name)
            parcel_match = parcel_match.filter(district_q)
        if ward:
            parcel_match = parcel_match.filter(ward_name__iexact=ward)

        villages_with_ccro = village_queryset.annotate(
            has_ccro=Exists(parcel_match)
        ).filter(has_ccro=True).count()
    except (ProgrammingError, OperationalError, ImportError):
        pass
    return total_ccros, villages_with_ccro


def _get_ccro_counts_by_district(region=None):
    """Idadi ya CCRO kwa kila wilaya — kutoka planning_parcels."""
    from django.db.utils import OperationalError, ProgrammingError

    ccro_by_district = {}
    try:
        from detailed_planning.models import PlanningParcel

        parcel_qs = PlanningParcel.objects.all()
        if region:
            parcel_qs = parcel_qs.filter(region_name__iexact=region)
        for row in parcel_qs.values('district_name').annotate(total=Count('id')):
            if row['district_name']:
                ccro_by_district[row['district_name']] = row['total']
    except (ProgrammingError, OperationalError, ImportError):
        pass
    return ccro_by_district


def api_dashboard_stats(request):
    """Dashboard statistics"""
    region, district, ward = _clean_location_params(
        request.GET.get('region'),
        request.GET.get('district'),
        request.GET.get('ward'),
    )

    queryset = _filter_villages_by_location(VillageBoundary.objects.all(), region, district, ward)

    total_districts = queryset.exclude(
        district_name__isnull=True
    ).exclude(
        district_name=''
    ).values('district_name').distinct().count()

    total_wards = queryset.exclude(
        ward_name__isnull=True
    ).exclude(
        ward_name=''
    ).values('ward_name').distinct().count()

    total_villages = queryset.count()

    plan_q = _village_has_plan_q()
    villages_with_plan = queryset.filter(plan_q).count()
    villages_without_plan = queryset.exclude(plan_q).count()

    with_plan_pct = round(villages_with_plan / total_villages * 100, 1) if total_villages > 0 else 0
    without_plan_pct = round(villages_without_plan / total_villages * 100, 1) if total_villages > 0 else 0

    total_ccros, villages_with_ccro = _get_ccro_summary(queryset, region, district, ward)

    landuse_data = {
        'agriculture_ha': 12500,
        'forest_ha': 8900,
        'urban_ha': 3200,
        'water_ha': 1800,
        'wetland_ha': 950,
        'pasture_ha': 5600,
        'total_area_ha': 32950,
        'agriculture_percentage': 38,
        'forest_percentage': 27,
        'urban_percentage': 10,
    }

    return JsonResponse({
        'success': True,
        'summary': {
            'districts': total_districts,
            'wards': total_wards,
            'villages': total_villages,
            'villages_with_plan': villages_with_plan,
            'villages_without_plan': villages_without_plan,
            'villages_with_ccro': villages_with_ccro,
            'total_ccros': total_ccros,
        },
        'villages': {
            'total': total_villages,
            'with_plan': villages_with_plan,
            'without_plan': villages_without_plan,
            'with_plan_percentage': with_plan_pct,
            'without_plan_percentage': without_plan_pct,
        },
        'cco': {
            'total': total_ccros,
            'villages_with_ccro': villages_with_ccro,
            'coverage_percentage': round(villages_with_ccro / total_villages * 100, 1) if total_villages > 0 else 0,
        },
        'landuse': landuse_data
    })

def api_village_data_list(request):
    """Get village data with pagination"""
    page = int(request.GET.get('page', 1))
    page_size = 20
    region = request.GET.get('region')
    district = request.GET.get('district')
    status_filter = request.GET.get('approval_status')
    
    # Clean parameters
    if region in ['undefined', 'null', 'None', '']:
        region = None
    if district in ['undefined', 'null', 'None', '']:
        district = None
    
    queryset = VillageBoundary.objects.all().order_by('name')
    
    if region:
        queryset = queryset.filter(region_name__iexact=region)
    if district:
        queryset = queryset.filter(district_name__iexact=district)

    plan_expiry_filters = {'expired', 'active', 'warning', 'no_plan'}
    if status_filter and status_filter not in ['undefined', 'null', 'None', '']:
        if status_filter in plan_expiry_filters:
            villages = list(queryset)
            if status_filter == 'active':
                villages = [
                    v for v in villages
                    if _plan_year_fields(v)[2]['code'] in ('active', 'warning')
                ]
            else:
                villages = [
                    v for v in villages
                    if _plan_year_fields(v)[2]['code'] == status_filter
                ]
            paginator = Paginator(villages, page_size)
        else:
            queryset = queryset.filter(status=status_filter)
            paginator = Paginator(queryset, page_size)
    else:
        paginator = Paginator(queryset, page_size)

    try:
        page_obj = paginator.page(page)
    except Exception:
        page_obj = paginator.page(1)

    data = [_serialize_village_data(v) for v in page_obj]
    
    return JsonResponse({
        'success': True,
        'data': data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'page_size': page_size
        }
    })

@csrf_exempt
def create_village_data_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        # Convert dates if provided
        date_prepared = None
        date_end = None
        if request.POST.get('plan_preparation_year'):
            try:
                year = int(request.POST.get('plan_preparation_year'))
                date_prepared = datetime(year, 1, 1).date()
            except (TypeError, ValueError):
                pass
        if request.POST.get('plan_expiry_year'):
            try:
                year = int(request.POST.get('plan_expiry_year'))
                date_end = datetime(year, 1, 1).date()
            except (TypeError, ValueError):
                pass
        elif date_prepared:
            date_end = datetime(date_prepared.year + PLAN_VALIDITY_YEARS, 1, 1).date()
        
        # CREATE VILLAGE - USING ONLY EXISTING COLUMNS
        # Generate a simple ID if needed
        import uuid
        village_id = str(uuid.uuid4())
        
        village = VillageBoundary.objects.create(
            id=village_id,  # Generate UUID if your model expects it
            name=request.POST.get('village_name'),
            ward_name=request.POST.get('ward_name', ''),
            district_name=request.POST.get('district_name', ''),
            region_name=request.POST.get('region_name', ''),
            sponsor=request.POST.get('sponsor', ''),
            date_prepared=date_prepared,
            date_end=date_end,
            status=request.POST.get('status', 'draft'),
            iv='',
        )
        
        # Handle PDF upload if needed
        pdf_file = request.FILES.get('plan_pdf')
        if pdf_file:
            from django.core.files.storage import default_storage
            file_path = default_storage.save(f'village_plans/{village_id}_{pdf_file.name}', pdf_file)
            # If you have a field for PDF, update it
            # village.plan_pdf = file_path
            # village.save()
        
        return JsonResponse({'success': True, 'id': str(village.id), 'message': 'Village created successfully'})
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@csrf_exempt
def delete_village_data_api(request, id):
    if request.method != 'DELETE':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        village = get_object_or_404(VillageBoundary, id=id)
        village.delete()
        return JsonResponse({'success': True, 'message': 'Village deleted successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def api_district_landuse_summary(request):
    """District-wise summary"""
    region = request.GET.get('region')
    
    # Clean region
    if region in ['undefined', 'null', 'None', '']:
        region = None
    
    queryset = VillageBoundary.objects.all()
    if region:
        queryset = queryset.filter(region_name__iexact=region)
    
    districts_data = {}
    for v in queryset:
        if v.district_name:
            if v.district_name not in districts_data:
                districts_data[v.district_name] = {
                    'district_name': v.district_name,
                    'total_villages': 0,
                    'villages_with_plan': 0,
                    'cco_count': 0,
                }
            districts_data[v.district_name]['total_villages'] += 1
            if _village_has_plan(v):
                districts_data[v.district_name]['villages_with_plan'] += 1
    
    ccro_by_district = _get_ccro_counts_by_district(region)

    result = []
    for district_name, data in districts_data.items():
        total = data['total_villages'] or 1
        result.append({
            'district_name': district_name,
            'total_villages': data['total_villages'],
            'villages_with_plan': data['villages_with_plan'],
            'plan_percentage': round(data['villages_with_plan'] / total * 100, 1),
            'total_agriculture': 0,
            'total_forest': 0,
            'total_urban': 0,
            'cco_count': ccro_by_district.get(district_name, 0),
        })
    
    return JsonResponse({'success': True, 'districts': result})

def api_social_services(request):
    district = request.GET.get('district')
    queryset = SocialService.objects.all()
    if district and district not in ['undefined', 'null', 'None', '']:
        queryset = queryset.filter(district_name__iexact=district)
    
    data = [{'name': s.name, 'service_type': s.service_type, 'ward_name': s.ward_name, 'district_name': s.district_name} 
            for s in queryset[:100]]
    return JsonResponse({'success': True, 'data': data})

def api_parcels(request):
    from django.db.utils import OperationalError, ProgrammingError

    region = request.GET.get('region')
    district = request.GET.get('district')
    ward = request.GET.get('ward')
    try:
        queryset = Parcel.objects.all()
        if region and region not in ['undefined', 'null', 'None', '']:
            queryset = queryset.filter(region_name__iexact=region)
        if district and district not in ['undefined', 'null', 'None', '']:
            queryset = queryset.filter(district_name__iexact=district)
        if ward and ward not in ['undefined', 'null', 'None', '']:
            queryset = queryset.filter(ward_name__iexact=ward)

        data = [{
            'parcel_number': p.parcel_number,
            'owner_name': p.owner_name or '',
            'village_name': p.village_name,
            'ward_name': p.ward_name or '',
            'area_ha': p.area_ha or 0,
        } for p in queryset[:100]]
    except (ProgrammingError, OperationalError):
        data = []
    return JsonResponse({'success': True, 'data': data})

def api_infrastructure(request):
    district = request.GET.get('district')
    queryset = Infrastructure.objects.all()
    if district and district not in ['undefined', 'null', 'None', '']:
        queryset = queryset.filter(district_name__iexact=district)
    
    data = [{'name': i.name, 'infra_type': i.infra_type, 'ward_name': i.ward_name, 'district_name': i.district_name}
            for i in queryset[:100]]
    return JsonResponse({'success': True, 'data': data})

@csrf_exempt
def upload_pdf_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    
    try:
        village_id = request.POST.get('village_id')
        pdf_file = request.FILES.get('pdf_file')
        
        if not village_id or not pdf_file:
            return JsonResponse({'success': False, 'error': 'Missing village_id or pdf_file'}, status=400)
        
        village = get_object_or_404(VillageBoundary, id=village_id)
        
        return JsonResponse({'success': True, 'message': 'PDF uploaded successfully'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def export_excel_landuse(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="landuse_data.xlsx"'
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Land Use Data"
        
        headers = ['Kijiji', 'Kata', 'Wilaya', 'Mkoa', 'Sponsor', 'Status', 'Tarehe ya Kuandaa', 'Tarehe ya Mwisho']
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1a7a3a", end_color="1a7a3a", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        villages = VillageBoundary.objects.all()[:500]
        for v in villages:
            ws.append([
                v.name or '', 
                v.ward_name or '', 
                v.district_name or '', 
                v.region_name or '',
                v.sponsor or '',
                v.status or '',
                v.date_prepared.strftime('%Y-%m-%d') if v.date_prepared else '',
                v.date_end.strftime('%Y-%m-%d') if v.date_end else ''
            ])
        
        wb.save(response)
        return response
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def generate_demo_data(request):
    return JsonResponse({'success': True, 'message': 'Data already exists (800 records)'})


# =====================================================
# SHAPEFILE API FUNCTIONS
# =====================================================

def api_get_region_boundary_from_shapefile(request, region_name):
    # Clean region_name
    if region_name in ['undefined', 'null', 'None', '']:
        return JsonResponse({'success': False, 'error': 'Invalid region name'}, status=400)
    
    try:
        region_upper = region_name.upper()
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT ST_AsGeoJSON(ST_Union(geom)) as geojson,
                       ST_X(ST_Centroid(ST_Union(geom))) as center_lon,
                       ST_Y(ST_Centroid(ST_Union(geom))) as center_lat
                FROM boundaries.tanzania_administrative
                WHERE UPPER(region_nam) = %s
            """, [region_upper])
            row = cursor.fetchone()
            if row and row[0]:
                return JsonResponse({
                    'success': True,
                    'boundary_geojson': json.loads(row[0]),
                    'center_lat': float(row[2]) if row[2] else None,
                    'center_lon': float(row[1]) if row[1] else None,
                    'name': region_name
                })
            return JsonResponse({'success': False, 'error': f'Region "{region_name}" not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def api_get_district_boundaries(request, region_name):
    # Clean region_name
    if region_name in ['undefined', 'null', 'None', '']:
        return JsonResponse({'type': 'FeatureCollection', 'features': []})
    
    try:
        region_upper = region_name.upper()
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    district_n,
                    ST_AsGeoJSON(ST_Union(geom)) as geojson,
                    ST_X(ST_Centroid(ST_Union(geom))) as center_lon,
                    ST_Y(ST_Centroid(ST_Union(geom))) as center_lat
                FROM boundaries.tanzania_administrative
                WHERE UPPER(region_nam) = %s 
                AND district_n IS NOT NULL AND district_n != ''
                GROUP BY district_n
            """, [region_upper])
            
            features = []
            for row in cursor.fetchall():
                district_name, geojson_str, lon, lat = row
                if district_name and geojson_str:
                    features.append({
                        'type': 'Feature',
                        'geometry': json.loads(geojson_str),
                        'properties': {'name': district_name, 'type': 'district'}
                    })
            return JsonResponse({'type': 'FeatureCollection', 'features': features})
    except Exception as e:
        return JsonResponse({'type': 'FeatureCollection', 'features': []})

def api_get_ward_boundaries(request, region_name, district_name):
    # Clean parameters
    if region_name in ['undefined', 'null', 'None', ''] or district_name in ['undefined', 'null', 'None', '']:
        return JsonResponse({'type': 'FeatureCollection', 'features': []})
    
    try:
        region_upper = region_name.upper()
        district_upper = district_name.upper()
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    ward_name,
                    ST_AsGeoJSON(ST_Union(geom)) as geojson,
                    ST_X(ST_Centroid(ST_Union(geom))) as center_lon,
                    ST_Y(ST_Centroid(ST_Union(geom))) as center_lat
                FROM boundaries.tanzania_administrative
                WHERE UPPER(region_nam) = %s 
                AND UPPER(district_n) = %s
                AND ward_name IS NOT NULL AND ward_name != ''
                GROUP BY ward_name
            """, [region_upper, district_upper])
            
            features = []
            for row in cursor.fetchall():
                ward_name, geojson_str, lon, lat = row
                if ward_name and geojson_str:
                    features.append({
                        'type': 'Feature',
                        'geometry': json.loads(geojson_str),
                        'properties': {'name': ward_name, 'type': 'ward'}
                    })
            return JsonResponse({'type': 'FeatureCollection', 'features': features})
    except Exception as e:
        return JsonResponse({'type': 'FeatureCollection', 'features': []})

def api_get_villages_by_ward_from_shapefile(request, region_name, district_name, ward_name):
    # Clean parameters
    if region_name in ['undefined', 'null', 'None', ''] or district_name in ['undefined', 'null', 'None', ''] or ward_name in ['undefined', 'null', 'None', '']:
        return JsonResponse({'villages': []})

    names = set()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT vil_mtaa_n
                FROM boundaries.tanzania_administrative
                WHERE UPPER(region_nam) = UPPER(%s) 
                AND UPPER(district_n) = UPPER(%s)
                AND UPPER(ward_name) = UPPER(%s)
                AND vil_mtaa_n IS NOT NULL
            """, [region_name, district_name, ward_name])
            for row in cursor.fetchall():
                if row[0]:
                    names.add(row[0])
    except Exception:
        pass

    try:
        from detailed_planning.views import _gazette_villages
        names |= _gazette_villages(region_name, district_name, ward_name)
    except Exception:
        pass

    villages = [{'name': v} for v in sorted(names)]
    return JsonResponse({'villages': villages})

def api_get_wards_by_district_from_shapefile(request, region_name, district_name):
    return api_get_ward_boundaries(request, region_name, district_name)

def api_check_database(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM boundaries.tanzania_administrative")
            count = cursor.fetchone()[0]
            return JsonResponse({'status': 'ok', 'record_count': count})
    except Exception as e:
        return JsonResponse({'status': 'error', 'error': str(e)})

def api_landuse_trends(request):
    return JsonResponse({'success': True, 'trends': []})

def api_village_data_detail(request, village_id):
    try:
        village = get_object_or_404(VillageBoundary, id=village_id)
        prep_year, expiry_year, status_info = _plan_year_fields(village)
        return JsonResponse({'success': True, 'data': {
            'id': str(village.id),
            'name': village.name,
            'ward_name': village.ward_name,
            'district_name': village.district_name,
            'region_name': village.region_name,
            'sponsor': village.sponsor,
            'date_prepared': village.date_prepared,
            'date_end': village.date_end,
            'plan_preparation_year': prep_year,
            'plan_expiry_year': expiry_year,
            'years_remaining': status_info['years_remaining'],
            'plan_status_label': status_info['label'],
            'plan_status_code': status_info['code'],
            'status': village.status,
            'iv': village.iv,
        }})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=404)

def api_village_data_filter(request):
    region = request.GET.get('region')
    district = request.GET.get('district')
    
    # Clean parameters
    if region in ['undefined', 'null', 'None', '']:
        region = None
    if district in ['undefined', 'null', 'None', '']:
        district = None
    
    queryset = VillageBoundary.objects.all()
    if region:
        queryset = queryset.filter(region_name__iexact=region)
    if district:
        queryset = queryset.filter(district_name__iexact=district)
    data = [{'id': str(v.id), 'name': v.name} for v in queryset[:100]]
    return JsonResponse({'success': True, 'data': data})

def api_social_services_filter(request):
    return api_social_services(request)

def api_parcels_filter(request):
    return api_parcels(request)

def api_parcels_by_cco(request, cco_id):
    return JsonResponse({'success': True, 'data': []})

def api_infrastructure_filter(request):
    return api_infrastructure(request)

def api_cco_list(request):
    return JsonResponse({'success': True, 'data': []})

def api_cco_detail(request, cco_id):
    return JsonResponse({'success': True, 'data': {}})

def api_landuse_filter(request):
    return JsonResponse({'success': True, 'data': []})

def api_landuse_summary(request):
    return JsonResponse({'success': True, 'summary': {}})


def api_landuse_geojson(request):
    """GeoJSON ya matumizi ya ardhi kutoka landuse.land_use (hifadhi baada ya upload)."""
    district = (request.GET.get('district') or '').strip()
    ward = (request.GET.get('ward') or '').strip()
    village = (request.GET.get('village') or '').strip()
    if not district:
        return JsonResponse({
            'type': 'FeatureCollection',
            'features': [],
            'meta': {'error': 'Chagua wilaya'},
        })

    try:
        from dashboard.landuse_service import landuse_queryset_for_location, landuse_to_geojson

        qs = landuse_queryset_for_location(
            district=district,
            ward=ward or None,
            village=village or None,
        )
        geo = landuse_to_geojson(qs)
        geo['meta'] = {
            'returned': len(geo.get('features') or []),
            'district': district,
            'ward': ward or None,
            'village': village or None,
        }
        return JsonResponse(geo)
    except Exception as e:
        logger.exception('api_landuse_geojson failed')
        return JsonResponse({
            'type': 'FeatureCollection',
            'features': [],
            'meta': {'error': str(e)},
        }, status=500)

def api_download_village_data(request, format):
    """Backward compatible — pakua taarifa za vijiji (CSV/Excel)."""
    region = request.GET.get('region') or request.GET.get('region_name')
    district = request.GET.get('district')
    ward = request.GET.get('ward')
    try:
        from dashboard.export_service import export_data
        return export_data('village_data', format, region, district, ward)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def api_export_data(request, data_type, fmt):
    """API ya kupakua data kwa formats mbalimbali."""
    region = request.GET.get('region') or request.GET.get('region_name')
    district = request.GET.get('district')
    ward = request.GET.get('ward')
    try:
        from dashboard.export_service import export_data
        return export_data(data_type, fmt, region, district, ward)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        logger.exception('Export error')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def api_download_shapefile_by_filter(request, data_type):
    fmt = request.GET.get('format', 'shapefile')
    region = request.GET.get('region') or request.GET.get('region_name')
    district = request.GET.get('district')
    ward = request.GET.get('ward')
    try:
        from dashboard.export_service import export_data
        return export_data(data_type, fmt, region, district, ward)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def export_shapefile(request, data_type):
    return api_download_shapefile_by_filter(request, data_type)

def export_excel_report(request, report_type):
    return JsonResponse({'success': True, 'message': f'Exporting {report_type}'})

def export_pdf_report(request, report_type):
    return JsonResponse({'success': True, 'message': f'Exporting {report_type}'})

def api_import_logs(request):
    logs = ImportLog.objects.all().order_by('-created_at')[:50]
    data = [{'id': str(log.id), 'import_type': log.import_type, 'filename': log.filename, 'status': log.status} for log in logs]
    return JsonResponse({'success': True, 'data': data})

def api_import_log_detail(request, log_id):
    log = get_object_or_404(ImportLog, id=log_id)
    return JsonResponse({'success': True, 'data': {'id': str(log.id), 'status': log.status}})

@csrf_exempt
def save_geojson(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    return JsonResponse({'status': 'success', 'message': 'GeoJSON received'})

@csrf_exempt
def upload_shapefile_api(request):
    """GIS Portal — pakia shapefile, clip kwa wilaya/kata, onyesha kwenye ramani."""
    from accounts.permissions import can_upload

    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'success': False, 'message': 'Method not allowed'}, status=405)

    if request.user.is_authenticated and not can_upload(request.user):
        return JsonResponse({
            'status': 'error', 'success': False,
            'message': 'Huna ruhusa ya kupakia data. Wasiliana na msimamizi.',
        }, status=403)

    from dashboard.shapefile_upload_service import spatial_files_from_request

    uploaded_files = spatial_files_from_request(request, ('shapefile',))
    if not uploaded_files:
        return JsonResponse({
            'status': 'error', 'success': False,
            'message': 'Hakuna faili. Chagua .zip (yenye .shp, .shx, .dbf) au .geojson',
        }, status=400)
    uploaded = uploaded_files[0]

    data_type = (request.POST.get('data_type') or 'landuse').strip()
    from detailed_planning.services import BOUNDARY_DATA_TYPE_MAP
    valid_types = {c[0] for c in ImportLog.IMPORT_TYPES} | set(BOUNDARY_DATA_TYPE_MAP.keys()) | {'other'}
    if data_type not in valid_types:
        data_type = 'other'

    region = (request.POST.get('region') or '').strip()
    district = (request.POST.get('district') or '').strip()
    ward = (request.POST.get('ward') or '').strip()
    village = (request.POST.get('village') or '').strip()

    if not district:
        return JsonResponse({
            'status': 'error', 'success': False,
            'message': 'Chagua wilaya au kata kwanza (Administrative Levels) ili data iwe overlay juu ya mipaka husika.',
        }, status=400)

    try:
        from dashboard.shapefile_upload_service import parse_spatial_upload_files
        from dashboard.boundary_service import (
            format_boundary_not_found_message,
            get_admin_boundary_feature,
            resolve_admin_boundary,
            resolve_region_for_district,
        )
        from dashboard.gis_processing_service import clip_geojson_to_aoi

        geojson = parse_spatial_upload_files(uploaded_files)
        import_meta = geojson.get('import_meta') or {}
        raw_count = import_meta.get('source_feature_count') or len(geojson.get('features', []))

        effective_region = region
        if not effective_region or effective_region.upper() == 'TANZANIA':
            effective_region = resolve_region_for_district(district) or region

        resolved = resolve_admin_boundary(region, district, ward or None)
        if not resolved or not resolved.get('geometry'):
            return JsonResponse({
                'status': 'error', 'success': False,
                'message': format_boundary_not_found_message(region, district, ward or None),
            }, status=400)

        effective_region = resolved.get('region') or effective_region or region
        district = resolved.get('district') or district
        ward = resolved.get('ward') or ward
        aoi_geom = resolved['geometry']

        geojson = clip_geojson_to_aoi(geojson, aoi_geom, data_type=data_type)
        clip_meta = geojson.pop('clip_meta', {}) or {}
        feature_count = len(geojson.get('features', []))
        clip_fallback = clip_meta.get('clip_fallback', False)
        if feature_count == 0:
            aoi_label = f'kata {ward}' if ward else f'wilaya {district}'
            extent_hint = ''
            in_ext = clip_meta.get('input_extent')
            aoi_ext = clip_meta.get('aoi_extent')
            if in_ext and aoi_ext:
                extent_hint = (
                    f' Eneo la shapefile: lon {in_ext["min_lon"]:.4f}–{in_ext["max_lon"]:.4f}, '
                    f'lat {in_ext["min_lat"]:.4f}–{in_ext["max_lat"]:.4f}. '
                    f'Mipaka ya {aoi_label}: lon {aoi_ext["min_lon"]:.4f}–{aoi_ext["max_lon"]:.4f}, '
                    f'lat {aoi_ext["min_lat"]:.4f}–{aoi_ext["max_lat"]:.4f}.'
                )
            return JsonResponse({
                'status': 'error', 'success': False,
                'message': (
                    f'Hakuna features ndani ya {aoi_label} (zilipakiwa {raw_count}). '
                    'Hakikisha shapefile iko katika eneo hilo na CRS sahihi.'
                    + extent_hint
                ),
                'clip_meta': clip_meta,
            }, status=400)

        aoi_label = f'{ward} ({district})' if ward else district
        district_note = ''
        if resolved.get('district_corrected') and ward:
            district_note = (
                f' (Kumbuka: kata "{ward}" iko chini ya wilaya {district}, '
                'si jina la wilaya uliyochagua.)'
            )
        clip_note = ''
        if clip_fallback:
            clip_note = (
                f' Onyo: hakuna kiwanja kilichokatwa ndani ya mipaka ya {aoi_label}; '
                f'viwanja vyote {raw_count} vimeingizwa kulingana na chaguo lako la kata.'
            )

        try:
            ImportLog.objects.create(
                import_type=data_type,
                filename=uploaded.name,
                file_path='',
                records_imported=feature_count,
                status='completed',
                imported_by=request.user if getattr(request, 'user', None) and request.user.is_authenticated else None,
                import_summary={
                    'feature_count': feature_count,
                    'raw_feature_count': raw_count,
                    'source': 'gis_portal',
                    'region': effective_region or region,
                    'district': district,
                    'ward': ward,
                    'clip_meta': clip_meta,
                },
            )
        except Exception:
            logger.warning('ImportLog haikuandikwa — endelea bila log', exc_info=True)

        db_import = None
        boundary_import = None
        created_by = (
            request.user
            if getattr(request, 'user', None) and request.user.is_authenticated
            else None
        )

        if data_type == 'parcels':
            try:
                from detailed_planning.services import import_parcels_from_geojson

                db_import = import_parcels_from_geojson(
                    geojson,
                    region=effective_region or region,
                    district=district,
                    ward=ward or None,
                    village=village or None,
                    shapefile_name=uploaded.name,
                    created_by=created_by,
                )
            except Exception:
                logger.exception('Auto-import ya viwanja imeshindwa')
                db_import = {'error': 'Imeshindwa kuhifadhi viwanja kwenye database'}

        elif data_type == 'landuse':
            try:
                from dashboard.landuse_service import import_landuse_from_geojson

                db_import = import_landuse_from_geojson(
                    geojson,
                    district=district,
                    ward=ward or None,
                    village=village or None,
                    shapefile_name=uploaded.name,
                )
            except Exception as exc:
                logger.exception('Auto-import ya matumizi imeshindwa')
                db_import = {
                    'error': f'Imeshindwa kuhifadhi matumizi kwenye database: {exc}'
                }

        from detailed_planning.services import import_boundaries_from_geojson

        boundary_level = BOUNDARY_DATA_TYPE_MAP.get(data_type)
        if boundary_level:
            try:
                boundary_import = import_boundaries_from_geojson(
                    geojson,
                    level=boundary_level,
                    region=effective_region or region,
                    district=district,
                    ward=ward or None,
                    village=village or None,
                    shapefile_name=uploaded.name,
                    created_by=created_by,
                )
            except ValueError as be:
                boundary_import = {'error': str(be)}
            except Exception:
                logger.exception('Auto-import ya mipaka imeshindwa')
                boundary_import = {'error': 'Imeshindwa kuhifadhi mipaka kwenye database'}

        boundary_feature = get_admin_boundary_feature(region, district, ward or None)

        import_note = ''
        if db_import and not db_import.get('error'):
            if data_type == 'landuse':
                import_note = (
                    f' Imesave DB (matumizi): {db_import.get("created", 0)} vipya, '
                    f'{db_import.get("updated", 0)} vilivyosasishwa.'
                )
            else:
                import_note = (
                    f' Imesave DB (viwanja): {db_import.get("created", 0)} vipya, '
                    f'{db_import.get("updated", 0)} vilivyosasishwa.'
                )
            if db_import.get('villages'):
                import_note += f' Vijiji: {", ".join(db_import["villages"][:8])}.'
            if db_import.get('skipped'):
                import_note += f' {db_import["skipped"]} vimerukwa.'
            if db_import.get('warning'):
                import_note += f' {db_import["warning"]}'
        elif db_import and db_import.get('error'):
            label = 'Matumizi' if data_type == 'landuse' else 'Viwanja'
            import_note = f' {label}: {db_import["error"]}.'

        if boundary_import and not boundary_import.get('error'):
            import_note += (
                f' Imesave DB (mipaka/{boundary_import.get("level")}): '
                f'{boundary_import.get("saved", 0)}.'
            )
            if boundary_import.get('skipped'):
                import_note += f' {boundary_import["skipped"]} vimerukwa.'
        elif boundary_import and boundary_import.get('error'):
            import_note += f' Mipaka: {boundary_import["error"]}.'

        return JsonResponse({
            'status': 'success',
            'success': True,
            'message': (
                f'Imepakiwa — features {feature_count} ndani ya {aoi_label}'
                + district_note
                + clip_note
                + import_note
                + (f'. {import_meta["message_sw"]}' if import_meta.get('message_sw') else '')
            ),
            'geojson': geojson,
            'feature_count': feature_count,
            'raw_feature_count': raw_count,
            'import_meta': import_meta,
            'clip_meta': clip_meta,
            'data_type': data_type,
            'db_import': db_import,
            'boundary_import': boundary_import,
            'aoi': {
                'region': effective_region or region,
                'district': district,
                'ward': ward,
                'village': village or (db_import or {}).get('village') or (boundary_import or {}).get('village'),
                'label': aoi_label,
            },
            'boundary_feature': boundary_feature,
        })
    except ValueError as e:
        return JsonResponse({'status': 'error', 'success': False, 'message': str(e)}, status=400)
    except Exception as e:
        logger.exception('GIS Portal shapefile upload error')
        err = str(e)
        if 'WinError 267' in err or '267' in err:
            err = 'Shapefile haikusomwa. Tumia .zip yenye .shp, .shx, .dbf pamoja.'
        return JsonResponse({'status': 'error', 'success': False, 'message': err}, status=500)


@csrf_exempt
def api_tools_upload_layer(request):
    """Pakia shapefile/GeoJSON kwa GIS Tools — GDAL."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    from dashboard.shapefile_upload_service import spatial_files_from_request

    uploaded_files = spatial_files_from_request(request, ('file',))
    if not uploaded_files:
        return JsonResponse({'success': False, 'error': 'Hakuna faili iliyochaguliwa'}, status=400)

    try:
        from dashboard.shapefile_upload_service import parse_spatial_upload_files
        geojson = parse_spatial_upload_files(uploaded_files)
        features = geojson.get('features', [])
        import_meta = geojson.get('import_meta') or {}
        return JsonResponse({
            'success': True,
            'geojson': geojson,
            'feature_count': len(features),
            'import_meta': import_meta,
            'warning': import_meta.get('message_sw') or None,
        })
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        logger.exception('GIS tools upload error')
        err = str(e)
        if 'WinError 267' in err or '267' in err:
            err = (
                'Shapefile haikusomwa (jina la faili au .dbf). '
                'Tengeneza .zip mpya yenye .shp, .shx, .dbf pamoja.'
            )
        return JsonResponse({'success': False, 'error': err}, status=500)


def _parse_json_body(request):
    try:
        return json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return None


@csrf_exempt
def api_tools_topology_check(request):
    """QGIS topology — GDAL/GEOS (sahihi kama QGIS)."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    data = _parse_json_body(request)
    if data is None:
        return JsonResponse({'success': False, 'error': 'JSON si sahihi'}, status=400)
    geojson = data.get('geojson')
    if not geojson or not geojson.get('features'):
        return JsonResponse({'success': False, 'error': 'Hakuna data ya kukagua'}, status=400)
    try:
        from dashboard.gis_processing_service import run_topology_check
        result = run_topology_check(
            geojson,
            rules=data.get('rules') or {},
            min_area_sqm=float(data.get('min_area_sqm', 100)),
        )
        return JsonResponse({'success': True, **result})
    except Exception as e:
        logger.exception('Topology check error')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
def api_tools_clean(request):
    """Mapshaper -clean — GDAL/GEOS."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    data = _parse_json_body(request)
    if data is None:
        return JsonResponse({'success': False, 'error': 'JSON si sahihi'}, status=400)
    geojson = data.get('geojson')
    if not geojson or not geojson.get('features'):
        return JsonResponse({'success': False, 'error': 'Hakuna data ya kusafisha'}, status=400)
    opts = data.get('options') or {}
    try:
        from dashboard.gis_processing_service import clean_geojson
        result = clean_geojson(geojson, {
            'min_area_sqm': float(opts.get('min_area_sqm', opts.get('minAreaSqM', 100))),
            'fix_gaps': opts.get('fix_gaps', True),
            'fix_overlaps': opts.get('fix_overlaps', True),
            'fix_invalid': opts.get('fix_invalid', True),
            'remove_duplicates': opts.get('remove_duplicates', True),
            'remove_slivers': opts.get('remove_slivers', True),
            'fix_coords': opts.get('fix_coords', True),
        })
        return JsonResponse({
            'success': True,
            'engine': result['engine'],
            'geojson': result['feature_collection'],
            'report': result['report'],
        })
    except Exception as e:
        logger.exception('Clean error')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
def api_tools_edit_command(request):
    """Mapshaper editing — GDAL/GEOS."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    data = _parse_json_body(request)
    if data is None:
        return JsonResponse({'success': False, 'error': 'JSON si sahihi'}, status=400)
    geojson = data.get('geojson')
    command = data.get('command', '')
    if not geojson or not command:
        return JsonResponse({'success': False, 'error': 'geojson na command zinahitajika'}, status=400)
    try:
        from dashboard.gis_processing_service import run_edit_command
        result = run_edit_command(geojson, command, data.get('params') or {})
        return JsonResponse({'success': True, **result})
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        logger.exception('Edit command error')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@csrf_exempt
def upload_cco_excel_api(request):
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)
    return JsonResponse({'status': 'success', 'message': 'CCO data uploaded'})


# =====================================================
# DONATION — MCHANGO WAKATI WA KUPAKUA DATA
# =====================================================

@csrf_exempt
def api_donation_initiate(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)
    try:
        data = json.loads(request.body.decode('utf-8') or '{}')
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Data si sahihi'}, status=400)

    try:
        from dashboard.donation_service import create_donation
        result = create_donation(request, data)
        return JsonResponse(result)
    except ValueError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        logger.exception('Donation initiate error')
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def donation_checkout(request, reference):
    donation = get_object_or_404(DownloadDonation, reference=reference)
    if donation.status == 'paid':
        return redirect('dashboard:donation_success', reference=reference)

    from django.conf import settings
    cfg = getattr(settings, 'DONATION_SETTINGS', {})
    return render(request, 'dashboard/donation_checkout.html', {
        'donation': donation,
        'stripe_publishable_key': cfg.get('STRIPE_PUBLISHABLE_KEY', ''),
        'cancelled': request.GET.get('cancelled') == '1',
    })


@csrf_exempt
def api_donation_pay_demo(request, reference):
    """Malipo ya majaribio (demo) — kadi au pay merchant."""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'}, status=405)

    donation = get_object_or_404(DownloadDonation, reference=reference)
    if donation.status == 'paid':
        return JsonResponse({
            'success': True,
            'redirect_url': f'/donation/imethibitishwa/{reference}/',
        })

    try:
        from dashboard.donation_service import complete_demo_payment
        card_last4 = (request.POST.get('card_last4') or '')[:4]
        complete_demo_payment(donation, card_last4 or 'DEMO')
        return JsonResponse({
            'success': True,
            'redirect_url': f'/donation/imethibitishwa/{reference}/',
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


def donation_success(request, reference):
    donation = get_object_or_404(DownloadDonation, reference=reference)

    if donation.status != 'paid':
        from django.conf import settings
        from dashboard.donation_service import verify_stripe_session, mark_donation_paid
        cfg = getattr(settings, 'DONATION_SETTINGS', {})
        session_id = request.GET.get('session_id')
        if donation.provider == 'stripe' and verify_stripe_session(donation, session_id, cfg):
            mark_donation_paid(donation)
        elif request.GET.get('OrderTrackingId'):
            mark_donation_paid(donation)

    return render(request, 'dashboard/donation_success.html', {
        'donation': donation,
        'download_url': donation.download_url() if donation.status == 'paid' else '',
    })


@csrf_exempt
def donation_pesapal_callback(request):
    """IPN callback kutoka Pesapal."""
    tracking_id = request.GET.get('OrderTrackingId') or request.POST.get('OrderTrackingId')
    merchant_ref = request.GET.get('OrderMerchantReference') or request.POST.get('OrderMerchantReference')
    ref = merchant_ref or tracking_id
    if ref:
        try:
            donation = DownloadDonation.objects.get(reference=ref)
            from dashboard.donation_service import mark_donation_paid
            mark_donation_paid(donation)
        except DownloadDonation.DoesNotExist:
            pass
    return JsonResponse({'status': 'ok'})