"""
Huduma ya kupakua data kwa formats mbalimbali (Shapefile, GeoJSON, GDB, n.k.)
"""
import csv
import io
import json
import logging
import os
import tempfile
import zipfile

from django.db import connection
from django.db.models import Q
from django.http import HttpResponse

from dashboard.models import VillageBoundary, SocialService, Parcel

logger = logging.getLogger(__name__)


SPATIAL_DATA_TYPES = {
    'region_boundary',
    'district_boundaries',
    'ward_boundaries',
    'village_boundaries',
}

TABULAR_DATA_TYPES = {
    'village_data',
    'landuse',
    'social_services',
    'parcels',
}

SPATIAL_FORMATS = {'shapefile', 'shp', 'geojson', 'json', 'geodatabase', 'gdb', 'gpkg', 'kml'}
TABULAR_FORMATS = {'csv', 'excel', 'xlsx'}
JSON_FORMATS = {'geojson', 'json'}


def _normalize_region(region):
    if not region:
        return None
    r = str(region).strip()
    if r.lower() in ('tanzania', 'all', 'undefined', 'null', 'none', ''):
        return None
    return r


def _normalize_fmt(fmt):
    fmt = (fmt or 'csv').strip().lower()
    if fmt == 'xlsx':
        return 'excel'
    if fmt == 'json':
        return 'geojson'
    return fmt


def _empty_fc():
    return {'type': 'FeatureCollection', 'features': []}


def _admin_table_exists() -> bool:
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT EXISTS (
                    SELECT 1
                    FROM information_schema.tables
                    WHERE table_schema = 'boundaries'
                      AND table_name = 'tanzania_administrative'
                )
                """
            )
            row = cursor.fetchone()
            return bool(row and row[0])
    except Exception:
        logger.exception('Failed checking boundaries.tanzania_administrative')
        return False


def _build_boundary_geojson(data_type, region=None, district=None, ward=None):
    """Tengeneza FeatureCollection kutoka boundaries.tanzania_administrative."""
    region = _normalize_region(region)
    district = (district or '').strip() or None
    ward = (ward or '').strip() or None

    where = ["geom IS NOT NULL"]
    params = []

    if region:
        where.append("UPPER(region_nam) = %s")
        params.append(region.upper())
    if district:
        where.append("UPPER(district_n) = %s")
        params.append(district.upper())
    if ward:
        where.append("UPPER(ward_name) = %s")
        params.append(ward.upper())

    if data_type == 'region_boundary':
        sql = f"""
            SELECT region_nam, NULL::text, NULL::text, NULL::text,
                   ST_AsGeoJSON(ST_Union(geom))
            FROM boundaries.tanzania_administrative
            WHERE {' AND '.join(where)}
            AND region_nam IS NOT NULL AND region_nam != ''
            GROUP BY region_nam
        """
    elif data_type == 'district_boundaries':
        sql = f"""
            SELECT district_n, region_nam, NULL::text, NULL::text,
                   ST_AsGeoJSON(ST_Union(geom))
            FROM boundaries.tanzania_administrative
            WHERE {' AND '.join(where)}
            AND district_n IS NOT NULL AND district_n != ''
            GROUP BY district_n, region_nam
        """
    elif data_type == 'ward_boundaries':
        sql = f"""
            SELECT ward_name, region_nam, district_n, NULL::text,
                   ST_AsGeoJSON(ST_Union(geom))
            FROM boundaries.tanzania_administrative
            WHERE {' AND '.join(where)}
            AND ward_name IS NOT NULL AND ward_name != ''
            GROUP BY ward_name, region_nam, district_n
        """
    elif data_type == 'village_boundaries':
        sql = f"""
            SELECT vil_mtaa_n, region_nam, district_n, ward_name,
                   ST_AsGeoJSON(ST_Union(geom))
            FROM boundaries.tanzania_administrative
            WHERE {' AND '.join(where)}
            AND vil_mtaa_n IS NOT NULL AND vil_mtaa_n != ''
            GROUP BY vil_mtaa_n, region_nam, district_n, ward_name
        """
    else:
        return _empty_fc()

    features = []
    with connection.cursor() as cursor:
        cursor.execute(sql, params)
        for name, reg, dist, wrd, geojson_str in cursor.fetchall():
            if not geojson_str:
                continue
            props = {'name': name or ''}
            if reg:
                props['region_nam'] = reg
            if dist:
                props['district_n'] = dist
            if wrd:
                props['ward_name'] = wrd
            if data_type == 'region_boundary':
                props['type'] = 'region'
            elif data_type == 'district_boundaries':
                props['type'] = 'district'
            elif data_type == 'ward_boundaries':
                props['type'] = 'ward'
            else:
                props['type'] = 'village'
            features.append({
                'type': 'Feature',
                'geometry': json.loads(geojson_str),
                'properties': props,
            })

    return {'type': 'FeatureCollection', 'features': features}


def _planning_boundary_geojson(data_type, region=None, district=None, ward=None):
    """Fallback: mipaka kutoka detailed_planning (Neon / bila tanzania_administrative)."""
    from detailed_planning.models import (
        DistrictPlanningBoundary,
        VillagePlanningBoundary,
        WardPlanningBoundary,
    )
    from detailed_planning.services import (
        _geom_to_wgs84_dict,
        boundaries_to_feature_collection,
    )
    from dashboard.boundary_service import _district_search_names

    region = _normalize_region(region)
    district = (district or '').strip() or None
    ward = (ward or '').strip() or None

    try:
        if data_type == 'region_boundary':
            if not region:
                return _empty_fc()
            qs = DistrictPlanningBoundary.objects.filter(
                region_name__iexact=region,
                geom__isnull=False,
            )
            merged = None
            for obj in qs.iterator():
                geom = obj.geom
                if not geom:
                    continue
                if geom.srid != 4326:
                    geom = geom.clone()
                    geom.transform(4326)
                merged = geom if merged is None else merged.union(geom)
            geometry = _geom_to_wgs84_dict(merged)
            if not geometry:
                return _empty_fc()
            return {
                'type': 'FeatureCollection',
                'features': [{
                    'type': 'Feature',
                    'geometry': geometry,
                    'properties': {'name': region, 'type': 'region', 'region_nam': region},
                }],
            }

        if data_type == 'district_boundaries':
            if not region:
                return _empty_fc()
            qs = DistrictPlanningBoundary.objects.filter(
                region_name__iexact=region,
                geom__isnull=False,
            ).order_by('district_name')
            if district:
                dist_q = Q()
                for name in _district_search_names(district):
                    dist_q |= Q(district_name__iexact=name)
                qs = qs.filter(dist_q)
            return boundaries_to_feature_collection(
                qs, name_attr='district_name', feature_type='district'
            )

        if data_type == 'ward_boundaries':
            if not region:
                return _empty_fc()
            qs = WardPlanningBoundary.objects.filter(
                region_name__iexact=region,
                geom__isnull=False,
            ).order_by('ward_name')
            if district:
                dist_q = Q()
                for name in _district_search_names(district):
                    dist_q |= Q(district_name__iexact=name)
                qs = qs.filter(dist_q)
            if ward:
                qs = qs.filter(ward_name__iexact=ward)
            return boundaries_to_feature_collection(
                qs, name_attr='ward_name', feature_type='ward'
            )

        if data_type == 'village_boundaries':
            if not region:
                return _empty_fc()
            qs = VillagePlanningBoundary.objects.filter(
                region_name__iexact=region,
                geom__isnull=False,
            ).order_by('village_name')
            if district:
                dist_q = Q()
                for name in _district_search_names(district):
                    dist_q |= Q(district_name__iexact=name)
                qs = qs.filter(dist_q)
            if ward:
                qs = qs.filter(ward_name__iexact=ward)
            return boundaries_to_feature_collection(
                qs, name_attr='village_name', feature_type='village'
            )
    except Exception:
        logger.exception('planning boundary export failed (%s)', data_type)
        return _empty_fc()

    return _empty_fc()


def _spatial_geojson(data_type, region=None, district=None, ward=None):
    """Mipaka: jaribu tanzania_administrative, kisha detailed_planning."""
    fc = _empty_fc()
    if _admin_table_exists():
        try:
            fc = _build_boundary_geojson(data_type, region, district, ward)
        except Exception:
            logger.exception('tanzania_administrative export failed; falling back')
            fc = _empty_fc()
    if not fc.get('features'):
        fc = _planning_boundary_geojson(data_type, region, district, ward)
    return fc


def _filter_village_queryset(region=None, district=None, ward=None):
    qs = VillageBoundary.objects.all()
    region = _normalize_region(region)
    if region:
        qs = qs.filter(region_name__iexact=region)
    if district:
        qs = qs.filter(district_name__iexact=district.strip())
    if ward:
        qs = qs.filter(ward_name__iexact=ward.strip())
    return qs


def _export_village_csv(region, district, ward):
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="taarifa_vijiji.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow([
        'Jina la Kijiji', 'Kata', 'Wilaya', 'Mkoa', 'Mfadhili', 'Hali',
        'Tarehe Iliyoandaliwa', 'Tarehe ya Mwisho',
    ])
    for v in _filter_village_queryset(region, district, ward):
        writer.writerow([
            v.name or '',
            v.ward_name or '',
            v.district_name or '',
            v.region_name or '',
            v.sponsor or '',
            v.status or '',
            v.date_prepared.strftime('%Y-%m-%d') if v.date_prepared else '',
            v.date_end.strftime('%Y-%m-%d') if v.date_end else '',
        ])
    return response


def _export_village_excel(region, district, ward):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Taarifa za Vijiji"
    headers = [
        'Jina la Kijiji', 'Kata', 'Wilaya', 'Mkoa', 'Mfadhili', 'Hali',
        'Tarehe Iliyoandaliwa', 'Tarehe ya Mwisho',
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a7a3a", end_color="1a7a3a", fill_type="solid")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for v in _filter_village_queryset(region, district, ward):
        ws.append([
            v.name or '',
            v.ward_name or '',
            v.district_name or '',
            v.region_name or '',
            v.sponsor or '',
            v.status or '',
            v.date_prepared.strftime('%Y-%m-%d') if v.date_prepared else '',
            v.date_end.strftime('%Y-%m-%d') if v.date_end else '',
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="taarifa_vijiji.xlsx"'
    return response


def _model_to_geojson(queryset, name_field='name'):
    features = []
    for obj in queryset:
        geom = getattr(obj, 'geom', None) or getattr(obj, 'geometry', None)
        if not geom:
            continue
        try:
            g = geom
            if getattr(g, 'srid', None) and g.srid != 4326:
                g = g.clone()
                g.transform(4326)
            gj = json.loads(g.geojson)
        except Exception:
            continue
        props = {}
        for f in obj._meta.fields:
            if f.name in ('geom', 'geom_point', 'geometry'):
                continue
            val = getattr(obj, f.name, None)
            if val is not None:
                props[f.name] = str(val) if not isinstance(val, (int, float, bool)) else val
        if name_field and getattr(obj, name_field, None):
            props['name'] = getattr(obj, name_field)
        features.append({'type': 'Feature', 'geometry': gj, 'properties': props})
    return {'type': 'FeatureCollection', 'features': features}


def _planning_parcels_geojson(region=None, district=None, ward=None, village=None):
    """Viwanja kutoka detailed_planning.planning_parcels (DB sahihi + WGS84)."""
    from detailed_planning.services import (
        _geom_to_wgs84_dict,
        parcels_for_location,
        serialize_ccro_shapefile_fields,
    )

    region = _normalize_region(region)
    district = (district or '').strip() or None
    ward = (ward or '').strip() or None
    village = (village or '').strip() or None

    features = []
    try:
        qs = parcels_for_location(region, district, ward, village).exclude(geom__isnull=True)
        for p in qs.iterator():
            geometry = _geom_to_wgs84_dict(p.geom)
            if not geometry:
                continue
            features.append({
                'type': 'Feature',
                'geometry': geometry,
                'properties': {
                    'parcel_number': p.parcel_number,
                    'is_identified': p.is_identified,
                    'owner_name': p.owner_name or '',
                    'region_name': p.region_name,
                    'district_name': p.district_name,
                    'ward_name': p.ward_name,
                    'village_name': p.village_name,
                    'area_ha': p.area_ha,
                    **serialize_ccro_shapefile_fields(p),
                },
            })
    except Exception:
        logger.exception('planning parcels export failed')
        return _empty_fc()
    return {'type': 'FeatureCollection', 'features': features}


def _get_tabular_as_geojson(data_type, region, district, ward, village=None):
    region = _normalize_region(region)
    q = Q()
    if region:
        q &= Q(region_name__iexact=region)
    if district:
        q &= Q(district_name__iexact=district.strip())
    if ward:
        q &= Q(ward_name__iexact=ward.strip())

    if data_type == 'parcels':
        # Priority: detailed planning parcels (actual Mpango Kinaa data)
        fc = _planning_parcels_geojson(region, district, ward, village)
        if fc.get('features'):
            return fc
        try:
            qs = Parcel.objects.filter(q) if q else Parcel.objects.all()
            return _model_to_geojson(qs, 'parcel_number')
        except Exception:
            return fc

    if data_type in ('village_data', 'landuse'):
        return _model_to_geojson(_filter_village_queryset(region, district, ward), 'name')

    if data_type == 'social_services':
        try:
            qs = SocialService.objects.filter(q) if q else SocialService.objects.all()
            return _model_to_geojson(qs, 'name')
        except Exception:
            return _empty_fc()

    return _empty_fc()


def _json_download_response(geojson_fc, base_name, fmt='geojson'):
    """Rudisha faili ya JSON/GeoJSON — bila GDAL, kamwe si HTML."""
    payload = geojson_fc if isinstance(geojson_fc, dict) else _empty_fc()
    if payload.get('type') != 'FeatureCollection':
        payload = {'type': 'FeatureCollection', 'features': payload.get('features') or []}
    body = json.dumps(payload, ensure_ascii=False).encode('utf-8')
    if fmt == 'json':
        content_type = 'application/json; charset=utf-8'
        filename = f'{base_name}.json'
    else:
        content_type = 'application/geo+json; charset=utf-8'
        filename = f'{base_name}.geojson'
    response = HttpResponse(body, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _ogr_spatial_export(geojson_fc, fmt, base_name):
    from osgeo import ogr, gdal

    gdal.UseExceptions()

    if not geojson_fc.get('features'):
        raise ValueError('Hakuna data ya kupakua kwa chujio ulilochagua.')

    geojson_bytes = json.dumps(geojson_fc, ensure_ascii=False).encode('utf-8')
    mem_path = f'/vsimem/{base_name}_export.geojson'
    gdal.FileFromMemBuffer(mem_path, geojson_bytes)

    try:
        if fmt == 'geojson':
            return geojson_bytes, 'application/geo+json', f'{base_name}.geojson'

        src = ogr.Open(mem_path)
        if src is None:
            raise ValueError('Imeshindwa kusoma data ya kijiografia.')

        with tempfile.TemporaryDirectory() as tmpdir:
            if fmt in ('shapefile', 'shp'):
                out_path = os.path.join(tmpdir, f'{base_name}.shp')
                drv = ogr.GetDriverByName('ESRI Shapefile')
                dst = drv.CopyDataSource(src, out_path)
                if dst:
                    dst = None
                src = None
                gdal.Unlink(mem_path)

                zip_path = os.path.join(tmpdir, f'{base_name}.zip')
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for fn in os.listdir(tmpdir):
                        if fn.endswith('.zip'):
                            continue
                        zf.write(os.path.join(tmpdir, fn), fn)
                with open(zip_path, 'rb') as f:
                    return f.read(), 'application/zip', f'{base_name}_shapefile.zip'

            if fmt in ('geodatabase', 'gdb'):
                # GDAL ya Windows kawaida haisaidii kuandika ESRI FileGDB — tumia GeoPackage
                gpkg_path = os.path.join(tmpdir, f'{base_name}.gpkg')
                drv = ogr.GetDriverByName('GPKG')
                dst = drv.CopyDataSource(src, gpkg_path)
                if dst:
                    dst = None
                src = None
                gdal.Unlink(mem_path)

                zip_path = os.path.join(tmpdir, f'{base_name}_geodatabase.zip')
                with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                    zf.write(gpkg_path, f'{base_name}.gpkg')
                    readme = (
                        'ESRI File Geodatabase (.gdb) haipatikani moja kwa moja kwenye seva hii.\n'
                        'Faili .gpkg ni GeoPackage — fungua kwa ArcGIS Pro, QGIS au GDAL.\n'
                    )
                    zf.writestr('README.txt', readme)
                with open(zip_path, 'rb') as f:
                    return f.read(), 'application/zip', f'{base_name}_geodatabase.zip'

            if fmt == 'gpkg':
                gpkg_path = os.path.join(tmpdir, f'{base_name}.gpkg')
                drv = ogr.GetDriverByName('GPKG')
                dst = drv.CopyDataSource(src, gpkg_path)
                if dst:
                    dst = None
                src = None
                gdal.Unlink(mem_path)
                with open(gpkg_path, 'rb') as f:
                    return f.read(), 'application/geopackage+sqlite3', f'{base_name}.gpkg'

            if fmt == 'kml':
                kml_path = os.path.join(tmpdir, f'{base_name}.kml')
                drv = ogr.GetDriverByName('KML')
                dst = drv.CopyDataSource(src, kml_path)
                if dst:
                    dst = None
                src = None
                gdal.Unlink(mem_path)
                with open(kml_path, 'rb') as f:
                    return f.read(), 'application/vnd.google-earth.kml+xml', f'{base_name}.kml'

        raise ValueError(f'Format "{fmt}" haitumiki.')
    finally:
        try:
            gdal.Unlink(mem_path)
        except Exception:
            pass


def _safe_base_name(data_type, region, district, ward):
    parts = [data_type]
    for label in (region, district, ward):
        if label:
            safe = ''.join(c if c.isalnum() or c in '-_' else '_' for c in str(label))
            parts.append(safe[:30])
    return '_'.join(parts)[:80] or 'export'


def export_data(data_type, fmt, region=None, district=None, ward=None, village=None):
    """
    Rudisha HttpResponse ya faili lililopakuliwa.
    Raises ValueError kwa makosa ya mtumiaji.
    """
    data_type = (data_type or 'village_data').strip().lower()
    requested_fmt = (fmt or 'csv').strip().lower()
    fmt = _normalize_fmt(fmt)

    if data_type in TABULAR_DATA_TYPES and fmt in TABULAR_FORMATS:
        if fmt == 'csv':
            return _export_village_csv(region, district, ward)
        if fmt == 'excel':
            return _export_village_excel(region, district, ward)

    if fmt in TABULAR_FORMATS and data_type in TABULAR_DATA_TYPES:
        raise ValueError('Chagua format ya kijiografia (Shapefile, GeoJSON, GDB, n.k.) kwa data hii.')

    if fmt not in SPATIAL_FORMATS and requested_fmt not in JSON_FORMATS:
        raise ValueError(f'Format "{requested_fmt}" haitumiki.')

    if data_type in SPATIAL_DATA_TYPES:
        geojson_fc = _spatial_geojson(data_type, region, district, ward)
    elif data_type in TABULAR_DATA_TYPES:
        geojson_fc = _get_tabular_as_geojson(data_type, region, district, ward, village)
    else:
        raise ValueError(f'Aina ya data "{data_type}" haijulikani.')

    base_name = _safe_base_name(data_type, region, district, ward)

    # JSON / GeoJSON: never depend on GDAL; always attachment
    if requested_fmt in JSON_FORMATS or fmt in JSON_FORMATS:
        out_fmt = 'json' if requested_fmt == 'json' else 'geojson'
        return _json_download_response(geojson_fc or _empty_fc(), base_name, out_fmt)

    if not (geojson_fc or {}).get('features'):
        raise ValueError('Hakuna data ya kupakua kwa chujio ulilochagua.')

    try:
        data, content_type, filename = _ogr_spatial_export(geojson_fc, fmt, base_name)
    except ImportError as exc:
        # GDAL missing — fall back to GeoJSON download instead of 500 HTML
        logger.warning('GDAL unavailable for %s export: %s — returning GeoJSON', fmt, exc)
        return _json_download_response(geojson_fc, base_name, 'geojson')
    except Exception:
        logger.exception('OGR export failed for %s/%s', data_type, fmt)
        raise

    response = HttpResponse(data, content_type=content_type)
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
